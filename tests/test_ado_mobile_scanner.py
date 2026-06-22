import json
import tempfile
import unittest
from datetime import datetime, timezone
from pathlib import Path
from unittest.mock import patch

import appsec_scan_router as scanner
import ado_mobile_scanner
import mobile_scanner
from openpyxl import load_workbook


def workbook_value(sheet, field_name, row):
    headers = [cell.value for cell in sheet[1]]
    return sheet.cell(row=row, column=headers.index(field_name) + 1).value


class PublicApiTests(unittest.TestCase):
    def test_package_api_is_importable(self):
        self.assertIs(scanner.ScanConfig, ado_mobile_scanner.ScanConfig)
        self.assertIs(scanner.ScanConfig, mobile_scanner.ScanConfig)
        self.assertTrue(callable(scanner.scan))
        self.assertTrue(callable(scanner.scan_to_reports))
        self.assertTrue(callable(scanner.detect_mobile_repo))
        self.assertTrue(callable(scanner.detect_inventory_repo))
        self.assertTrue(callable(scanner.AppSecInventoryService))
        self.assertTrue(callable(scanner.AppSecScanRouter))
        self.assertTrue(callable(scanner.GitHubEnterpriseClient))


class ProviderClientTests(unittest.TestCase):
    def test_normalizes_github_enterprise_api_urls(self):
        self.assertEqual(scanner.normalize_github_api_url(""), "https://api.github.com")
        self.assertEqual(
            scanner.normalize_github_api_url("https://github.fabrikam.example"),
            "https://github.fabrikam.example/api/v3",
        )
        self.assertEqual(
            scanner.normalize_github_api_url("https://github.fabrikam.example/api/v3"),
            "https://github.fabrikam.example/api/v3",
        )

    def test_normalizes_github_commits_for_activity_extraction(self):
        commit = scanner.github_commit_to_activity_commit(
            {
                "commit": {
                    "author": {
                        "name": "Alice Adams",
                        "email": "alice@example.com",
                        "date": "2026-06-01T12:00:00Z",
                    },
                    "committer": {
                        "name": "Build Service",
                        "email": "build@example.com",
                        "date": "2026-06-02T12:00:00Z",
                    },
                }
            }
        )

        self.assertEqual(commit["author"]["name"], "Alice Adams")
        self.assertEqual(commit["author"]["email"], "alice@example.com")
        self.assertEqual(commit["committer"]["date"], "2026-06-02T12:00:00Z")

    def test_parse_args_supports_github_enterprise(self):
        with patch.dict("os.environ", {"GITHUB_TOKEN": "token"}, clear=False):
            config = scanner.parse_args(
                [
                    "--provider",
                    "github-enterprise",
                    "--base-url",
                    "https://github.fabrikam.example/api/v3",
                    "--org",
                    "FabrikamCloud",
                    "--repo",
                    "mobile-app",
                    "--out-dir",
                    "reports",
                ]
            )

        self.assertEqual(config.provider, "github-enterprise")
        self.assertEqual(config.base_url, "https://github.fabrikam.example/api/v3")
        self.assertEqual(config.org, "FabrikamCloud")
        self.assertEqual(config.project, "mobile-app")
        self.assertEqual(config.pat, "token")

    def test_create_source_client_supports_github_enterprise(self):
        config = scanner.ScanConfig(
            org="FabrikamCloud",
            pat="token",
            project="mobile-app",
            out_dir=Path("reports"),
            out_prefix="scan",
            max_workers=1,
            content_workers=1,
            max_commits_per_repo=0,
            timeout_seconds=30,
            min_confidence="low",
            provider="github-enterprise",
            base_url="https://github.fabrikam.example",
        )

        client = scanner.create_source_client(config)
        try:
            self.assertIsInstance(client, scanner.GitHubEnterpriseClient)
            self.assertEqual(client.base_url, "https://github.fabrikam.example/api/v3")
        finally:
            client.close()


class UiServiceTests(unittest.TestCase):
    def test_normalize_scan_config_requires_org(self):
        with self.assertRaises(ValueError):
            scanner.normalize_scan_config({"provider": "azure-devops"})

    def test_normalize_scan_config_requires_github_base_url(self):
        with self.assertRaises(ValueError):
            scanner.normalize_scan_config(
                {
                    "provider": "github-enterprise",
                    "org": "FabrikamCloud",
                }
            )

    def test_build_scan_command_for_github_enterprise(self):
        config = scanner.normalize_scan_config(
            {
                "provider": "github-enterprise",
                "org": "FabrikamCloud",
                "repo": "mobile-app",
                "baseUrl": "https://github.fabrikam.example/api/v3",
                "outPrefix": "inventory scan",
                "minConfidence": "medium",
                "activityMode": "latest",
                "storeLookup": True,
            }
        )

        command = scanner.build_scan_command(config, Path("/reports/scan-1"))

        self.assertIn("--provider", command)
        self.assertIn("github-enterprise", command)
        self.assertIn("--base-url", command)
        self.assertIn("https://github.fabrikam.example/api/v3", command)
        self.assertIn("--repo", command)
        self.assertIn("mobile-app", command)
        self.assertIn("--store-lookup", command)
        self.assertIn("inventory_scan", command)

    def test_build_scan_command_scans_all_github_repos_when_repo_empty(self):
        config = scanner.normalize_scan_config(
            {
                "provider": "github-enterprise",
                "org": "FabrikamCloud",
                "baseUrl": "https://github.fabrikam.example/api/v3",
            }
        )

        command = scanner.build_scan_command(config, Path("/reports/scan-1"))

        self.assertNotIn("--repo", command)

    def test_build_scan_command_for_azure_devops(self):
        config = scanner.normalize_scan_config(
            {
                "provider": "azure-devops",
                "org": "FabrikamCloud",
                "project": "Go_To_Market",
                "activityMode": "contributors",
                "maxWorkers": 4,
            }
        )

        command = scanner.build_scan_command(config, Path("/reports/scan-2"))

        self.assertIn("--provider", command)
        self.assertIn("azure-devops", command)
        self.assertIn("--project", command)
        self.assertIn("Go_To_Market", command)
        self.assertNotIn("--base-url", command)
        self.assertIn("4", command)

    def test_build_scan_command_scans_all_ado_projects_when_project_empty(self):
        config = scanner.normalize_scan_config(
            {
                "provider": "azure-devops",
                "org": "FabrikamCloud",
            }
        )

        command = scanner.build_scan_command(config, Path("/reports/scan-2"))

        self.assertNotIn("--project", command)

    def test_redact_command_hides_pat_values(self):
        command = ("appsec-scan-router", "--pat", "secret", "--org", "FabrikamCloud")

        redacted = scanner.redact_command(command)

        self.assertEqual(redacted, ["appsec-scan-router", "--pat", "[redacted]", "--org", "FabrikamCloud"])


class DetectionTests(unittest.TestCase):
    def test_detects_react_native_android_repo(self):
        paths = [
            "/package.json",
            "/android/app/build.gradle",
            "/android/app/src/main/AndroidManifest.xml",
            "/metro.config.js",
        ]
        contents = {
            "/package.json": '{"dependencies": {"react-native": "0.75.0"}}',
            "/android/app/build.gradle": """\
plugins {
    id 'com.android.application'
}

android {
    namespace 'com.fabrikam.agsnap'
    defaultConfig {
        applicationId 'com.fabrikam.agsnap'
        versionName '1.0.2'
    }
}
""",
            "/android/app/src/main/AndroidManifest.xml": """\
<manifest xmlns:android="http://schemas.android.com/apk/res/android"
    package="com.fabrikam.agsnap" />
""",
        }

        confidence, evidence, score = scanner.detect_mobile_repo(paths, contents)
        categories = {item.category for item in evidence}

        self.assertEqual(confidence, "high")
        self.assertGreaterEqual(score, 7)
        self.assertIn("android", categories)
        self.assertIn("react_native", categories)

    def test_returns_none_when_no_mobile_signals_exist(self):
        confidence, evidence, score = scanner.detect_mobile_repo(
            ["/README.md", "/src/server.py"],
            {},
        )

        self.assertEqual(confidence, "none")
        self.assertEqual(evidence, [])
        self.assertEqual(score, 0)

    def test_detects_web_frontend_inventory_repo(self):
        confidence, evidence, score = scanner.detect_inventory_repo(
            ["/package.json", "/src/App.tsx"],
            {
                "/package.json": json.dumps(
                    {
                        "name": "customer-portal",
                        "version": "2.4.1",
                        "dependencies": {"react": "18.3.1", "vite": "5.0.0"},
                        "scripts": {"start": "vite"},
                    }
                )
            },
        )

        categories = {item.category for item in evidence}
        self.assertEqual(confidence, "medium")
        self.assertGreaterEqual(score, 4)
        self.assertIn("web_frontend", categories)

    def test_detects_spring_microservice_inventory_repo(self):
        confidence, evidence, score = scanner.detect_inventory_repo(
            ["/pom.xml", "/Dockerfile"],
            {
                "/pom.xml": """\
<project>
  <artifactId>orders-api</artifactId>
  <dependencies>
    <dependency><artifactId>spring-boot-starter-web</artifactId></dependency>
    <dependency><artifactId>spring-kafka</artifactId></dependency>
  </dependencies>
</project>
""",
                "/Dockerfile": "FROM eclipse-temurin:21\nCMD [\"java\", \"-jar\", \"app.jar\"]\n",
            },
        )

        categories = {item.category for item in evidence}
        self.assertEqual(confidence, "high")
        self.assertGreaterEqual(score, 8)
        self.assertIn("microservice", categories)
        self.assertIn("api_service", categories)
        self.assertIn("middleware", categories)
        self.assertIn("containerized_service", categories)

    def test_detects_python_middleware_inventory_repo(self):
        confidence, evidence, score = scanner.detect_inventory_repo(
            ["/pyproject.toml"],
            {
                "/pyproject.toml": """\
[project]
name = "billing-worker"
version = "0.8.0"
dependencies = ["celery", "confluent-kafka"]
"""
            },
        )

        categories = {item.category for item in evidence}
        self.assertEqual(confidence, "medium")
        self.assertGreaterEqual(score, 3)
        self.assertIn("middleware", categories)

    def test_generic_config_xml_is_not_mobile(self):
        confidence, evidence, score = scanner.detect_mobile_repo(
            ["/config.xml"],
            {"/config.xml": "<configuration><setting name='example' /></configuration>"},
        )

        self.assertEqual(confidence, "none")
        self.assertEqual(evidence, [])
        self.assertEqual(score, 0)

    def test_generic_csproj_is_not_xamarin_or_maui(self):
        confidence, evidence, score = scanner.detect_mobile_repo(
            ["/src/Api/Api.csproj"],
            {
                "/src/Api/Api.csproj": """\
<Project Sdk="Microsoft.NET.Sdk.Web">
  <PropertyGroup>
    <TargetFramework>net8.0</TargetFramework>
  </PropertyGroup>
</Project>
"""
            },
        )

        self.assertEqual(confidence, "none")
        self.assertEqual(evidence, [])
        self.assertEqual(score, 0)

    def test_resolved_gradle_app_id_contributes_to_detection_evidence(self):
        confidence, evidence, score = scanner.detect_mobile_repo(
            ["/gradle.properties", "/android/app/build.gradle"],
            {
                "/gradle.properties": "appId=com.fabrikam.agsnap\n",
                "/android/app/build.gradle": """\
plugins {
    id 'com.android.application'
}
android {
    defaultConfig {
        applicationId "${appId}"
    }
}
""",
            },
        )

        details = {item.detail for item in evidence}
        self.assertEqual(confidence, "high")
        self.assertGreaterEqual(score, 7)
        self.assertIn("Gradle applicationId com.fabrikam.agsnap", details)

    def test_should_fetch_allowed_content_files(self):
        self.assertTrue(scanner.should_fetch_content("/src/MyApp.csproj"))
        self.assertTrue(scanner.should_fetch_content("/package.json"))
        self.assertTrue(scanner.should_fetch_content("/pom.xml"))
        self.assertTrue(scanner.should_fetch_content("/pyproject.toml"))
        self.assertTrue(scanner.should_fetch_content("/Dockerfile"))
        self.assertTrue(scanner.should_fetch_content("/serverless.yml"))
        self.assertTrue(scanner.should_fetch_content("/gradle.properties"))
        self.assertTrue(scanner.should_fetch_content("/Directory.Build.props"))
        self.assertTrue(scanner.should_fetch_content("/android/app/src/main/AndroidManifest.xml"))
        self.assertTrue(scanner.should_fetch_content("/ios/App/Info.plist"))
        self.assertTrue(scanner.should_fetch_content("/android/app/src/main/res/values/strings.xml"))
        self.assertFalse(scanner.should_fetch_content("/src/app.py"))

    def test_normalize_path_adds_leading_slash_and_unix_separators(self):
        self.assertEqual(scanner.normalize_path("android\\app\\build.gradle"), "/android/app/build.gradle")


class MetadataExtractionTests(unittest.TestCase):
    def test_extracts_android_name_version_and_identifier(self):
        metadata = scanner.extract_mobile_metadata(
            {
                "/android/app/src/main/AndroidManifest.xml": """\
<manifest xmlns:android="http://schemas.android.com/apk/res/android"
    package="com.fabrikam.agsnap"
    android:versionName="1.0.2">
    <application android:label="@string/app_name" />
</manifest>
""",
                "/android/app/src/main/res/values/strings.xml": """\
<resources>
    <string name="app_name">Agsnap</string>
</resources>
""",
            }
        )

        self.assertEqual(metadata.name, "Agsnap")
        self.assertEqual(metadata.version, "1.0.2")
        self.assertEqual(metadata.identifier, "com.fabrikam.agsnap")

    def test_extracts_gradle_identifier_and_version(self):
        metadata = scanner.extract_mobile_metadata(
            {
                "/android/app/build.gradle.kts": """\
android {
    namespace = "com.fabrikam.agsnap"
    defaultConfig {
        applicationId = "com.fabrikam.agsnap"
        versionName = "1.0.2"
    }
}
"""
            }
        )

        self.assertEqual(metadata.version, "1.0.2")
        self.assertEqual(metadata.identifier, "com.fabrikam.agsnap")
        self.assertEqual(metadata.identifier_source, "Gradle applicationId/namespace")

    def test_resolves_gradle_identifier_from_properties(self):
        metadata = scanner.extract_mobile_metadata(
            {
                "/gradle.properties": "appId=com.fabrikam.agsnap\n",
                "/android/app/build.gradle": """\
android {
    defaultConfig {
        applicationId "${appId}"
    }
}
""",
            }
        )

        self.assertEqual(metadata.identifier, "com.fabrikam.agsnap")
        self.assertEqual(metadata.identifier_source, "Gradle applicationId/namespace")

    def test_extracts_ios_info_plist_metadata(self):
        metadata = scanner.extract_mobile_metadata(
            {
                "/ios/App/Info.plist": """\
<?xml version="1.0" encoding="UTF-8"?>
<plist version="1.0">
<dict>
  <key>CFBundleDisplayName</key><string>Agsnap</string>
  <key>CFBundleShortVersionString</key><string>1.0.2</string>
  <key>CFBundleIdentifier</key><string>com.fabrikam.agsnap</string>
</dict>
</plist>
"""
            }
        )

        self.assertEqual(metadata.name, "Agsnap")
        self.assertEqual(metadata.version, "1.0.2")
        self.assertEqual(metadata.identifier, "com.fabrikam.agsnap")

    def test_extracts_ios_metadata_from_xcode_settings(self):
        metadata = scanner.extract_mobile_metadata(
            {
                "/ios/App.xcodeproj/project.pbxproj": """\
PRODUCT_NAME = Agsnap;
MARKETING_VERSION = 1.0.2;
PRODUCT_BUNDLE_IDENTIFIER = com.fabrikam.agsnap;
"""
            }
        )

        self.assertEqual(metadata.name, "Agsnap")
        self.assertEqual(metadata.version, "1.0.2")
        self.assertEqual(metadata.identifier, "com.fabrikam.agsnap")
        self.assertEqual(metadata.identifier_source, "Xcode build settings")

    def test_resolves_ios_plist_identifier_from_xcode_settings(self):
        metadata = scanner.extract_mobile_metadata(
            {
                "/ios/App/Info.plist": """\
<plist version="1.0">
<dict>
  <key>CFBundleIdentifier</key><string>$(PRODUCT_BUNDLE_IDENTIFIER)</string>
</dict>
</plist>
""",
                "/ios/App.xcodeproj/project.pbxproj": """\
PRODUCT_BUNDLE_IDENTIFIER = $(PRODUCT_BUNDLE_IDENTIFIER);
PRODUCT_BUNDLE_IDENTIFIER = com.fabrikam.agsnap;
""",
            }
        )

        self.assertEqual(metadata.identifier, "com.fabrikam.agsnap")
        self.assertEqual(metadata.identifier_source, "Info.plist")

    def test_filters_placeholder_versions(self):
        metadata = scanner.extract_mobile_metadata(
            {
                "/package.json": json.dumps(
                    {
                        "name": "agsnap",
                        "version": "999.999.999",
                    }
                )
            }
        )

        self.assertEqual(metadata.name, "agsnap")
        self.assertEqual(metadata.version, "")

    def test_extracts_expo_metadata(self):
        metadata = scanner.extract_mobile_metadata(
            {
                "/app.json": json.dumps(
                    {
                        "expo": {
                            "name": "Agsnap",
                            "version": "1.0.2",
                            "ios": {"bundleIdentifier": "com.fabrikam.agsnap"},
                            "android": {"package": "com.fabrikam.agsnap"},
                        }
                    }
                )
            }
        )

        self.assertEqual(metadata.name, "Agsnap")
        self.assertEqual(metadata.version, "1.0.2")
        self.assertEqual(metadata.identifier, "com.fabrikam.agsnap")

    def test_extracts_capacitor_metadata(self):
        metadata = scanner.extract_mobile_metadata(
            {
                "/capacitor.config.ts": """\
export default {
  appId: 'com.fabrikam.agsnap',
  appName: 'Agsnap',
  version: '1.0.2'
}
"""
            }
        )

        self.assertEqual(metadata.name, "Agsnap")
        self.assertEqual(metadata.version, "1.0.2")
        self.assertEqual(metadata.identifier, "com.fabrikam.agsnap")


class RepoActivityTests(unittest.TestCase):
    def test_extracts_contributors_and_last_updated(self):
        activity = scanner.extract_repo_activity(
            [
                {
                    "author": {"name": "Alice Adams", "email": "alice@example.com"},
                    "committer": {
                        "name": "Build Service",
                        "email": "build@example.com",
                        "date": "2024-04-01T12:00:00Z",
                    },
                },
                {
                    "author": {"name": "Bob Brown", "email": "bob@example.com"},
                    "committer": {"name": "Bob Brown", "email": "bob@example.com", "date": "2024-05-02T08:30:15.123Z"},
                },
                {
                    "author": {"name": "Alice Adams", "email": "alice@example.com"},
                    "committer": {"name": "Alice Adams", "email": "alice@example.com", "date": "2024-03-01T00:00:00Z"},
                },
            ]
        )

        self.assertEqual(
            activity.contributing_developers,
            (
                "Alice Adams <alice@example.com>",
                "Bob Brown <bob@example.com>",
            ),
        )
        self.assertEqual(activity.last_updated, "2024-05-02T08:30:15Z")

    def test_fetch_repo_activity_latest_mode_only_requests_latest_commit(self):
        class FakeClient:
            def __init__(self):
                self.calls = []

            def list_commits(self, **kwargs):
                self.calls.append(kwargs)
                return [
                    {
                        "author": {"name": "Alice Adams", "email": "alice@example.com"},
                        "committer": {"date": "2024-05-02T08:30:15Z"},
                    }
                ]

        client = FakeClient()

        activity = scanner.fetch_repo_activity(
            client=client,
            project_name="Project",
            repo_id="repo-id",
            branch_name="main",
            max_commits=0,
            activity_mode="latest",
        )

        self.assertEqual(client.calls[0]["max_commits"], 1)
        self.assertEqual(activity.contributing_developers, ())
        self.assertEqual(activity.last_updated, "2024-05-02T08:30:15Z")

    def test_fetch_repo_activity_contributors_mode_uses_requested_commit_limit(self):
        class FakeClient:
            def __init__(self):
                self.calls = []

            def list_commits(self, **kwargs):
                self.calls.append(kwargs)
                return []

        client = FakeClient()

        scanner.fetch_repo_activity(
            client=client,
            project_name="Project",
            repo_id="repo-id",
            branch_name="main",
            max_commits=250,
            activity_mode="contributors",
        )

        self.assertEqual(client.calls[0]["max_commits"], 250)


class OutputTests(unittest.TestCase):
    class FakeBranchClient:
        def __init__(self, refs=None, definitions=None):
            self.refs = refs or []
            self.definitions = definitions or []
            self.branch_calls = 0
            self.definition_calls = 0

        def list_branches(self, project_name, repo_id):
            self.branch_calls += 1
            return self.refs

        def list_build_definitions_for_repo(self, project_name, repo_id):
            self.definition_calls += 1
            return self.definitions

    def sample_result(self):
        return {
            "project": "Project",
            "repo_name": "Repo",
            "branch_name": "main",
            "branch_last_updated": "2024-05-02T08:30:15Z",
            "branch_age_bucket": scanner.ACTIVE_SHEET_NAME,
            "web_url": "https://example.invalid/repo",
            "source_url": "https://example.invalid/repo.git",
            "inventory_name": "Agsnap",
            "inventory_version": "1.0.2",
            "inventory_types": "mobile_app",
            "primary_language": "Java/Kotlin",
            "scanner_target": "https://example.invalid/repo.git#branch=main",
            "semgrep_target": "https://example.invalid/repo.git#branch=main",
            "sonarqube_project_key": "Project:Repo:main",
            "sonarqube_project_name": "Agsnap",
            "mobile_name": "Agsnap",
            "mobile_version": "1.0.2",
            "mobile_identifier": "com.fabrikam.agsnap",
            "mobile_identifier_source": "Gradle applicationId/namespace",
            "mobile_identifier_status": "found",
            "contributing_developers": "Alice Adams <alice@example.com>; Bob Brown <bob@example.com>",
            "last_updated": "2024-05-02T08:30:15Z",
            "confidence": "medium",
            "score": 2,
            "categories": "android",
            **scanner.type_columns(["mobile_app"]),
            **scanner.category_columns(["android"]),
            "detection_evidence": json.dumps(
                [
                    {
                        "category": "android",
                        "source": "/android/app/build.gradle",
                        "detail": "Gradle applicationId com.fabrikam.agsnap",
                        "weight": 3,
                    }
                ]
            ),
        }

    def test_write_outputs_creates_csv_and_json(self):
        result = self.sample_result()

        with tempfile.TemporaryDirectory() as tmpdir:
            csv_path, json_path, xlsx_path = scanner.write_outputs([result], Path(tmpdir), "scan")

            self.assertTrue(csv_path.exists())
            self.assertTrue(json_path.exists())
            self.assertTrue(xlsx_path.exists())
            self.assertTrue((Path(tmpdir) / "scan_scanner_targets.csv").exists())
            self.assertTrue((Path(tmpdir) / "scan_scanner_targets.json").exists())
            self.assertTrue((Path(tmpdir) / "scan_semgrep_targets.txt").exists())
            self.assertTrue((Path(tmpdir) / "scan_sonarqube_projects.csv").exists())
            loaded = json.loads(json_path.read_text(encoding="utf-8"))
            self.assertEqual(loaded, [result])
            self.assertIn("repo_name", csv_path.read_text(encoding="utf-8"))
            self.assertIn("https://example.invalid/repo.git#branch=main", (Path(tmpdir) / "scan_semgrep_targets.txt").read_text(encoding="utf-8"))
            targets = json.loads((Path(tmpdir) / "scan_scanner_targets.json").read_text(encoding="utf-8"))
            self.assertEqual(targets[0]["inventory_name"], "Agsnap")
            workbook = load_workbook(xlsx_path)
            self.assertEqual(workbook.sheetnames, [scanner.ACTIVE_SHEET_NAME, scanner.OLDER_SHEET_NAME])
            self.assertEqual(workbook_value(workbook[scanner.ACTIVE_SHEET_NAME], "mobile_name", 2), "Agsnap")

    def test_streaming_report_writer_flushes_rows_as_they_are_written(self):
        result = self.sample_result()

        with tempfile.TemporaryDirectory() as tmpdir:
            with scanner.StreamingReportWriter(Path(tmpdir), "scan") as writer:
                self.assertTrue(writer.csv_path.exists())
                self.assertTrue(writer.json_path.exists())
                self.assertTrue(writer.xlsx_path.exists())
                self.assertTrue(writer.scanner_targets_csv_path.exists())
                self.assertTrue(writer.scanner_targets_json_path.exists())
                self.assertTrue(writer.semgrep_targets_path.exists())
                self.assertTrue(writer.sonarqube_projects_path.exists())

                writer.write_result(result)

                csv_text = writer.csv_path.read_text(encoding="utf-8")
                json_text = writer.json_path.read_text(encoding="utf-8")
                self.assertIn("Agsnap", csv_text)
                self.assertIn("Agsnap", json_text)

            loaded = json.loads(writer.json_path.read_text(encoding="utf-8"))
            self.assertEqual(loaded, [result])
            workbook = load_workbook(writer.xlsx_path)
            self.assertEqual(workbook_value(workbook[scanner.ACTIVE_SHEET_NAME], "mobile_name", 2), "Agsnap")

    def test_category_columns_are_excel_filter_friendly(self):
        columns = scanner.category_columns(["android", "react_native"])

        self.assertEqual(columns["category_android"], "TRUE")
        self.assertEqual(columns["category_react_native"], "TRUE")
        self.assertEqual(columns["category_ios"], "FALSE")

    def test_type_columns_are_excel_filter_friendly(self):
        columns = scanner.type_columns(["web_app", "microservice"])

        self.assertEqual(columns["type_web_app"], "TRUE")
        self.assertEqual(columns["type_microservice"], "TRUE")
        self.assertEqual(columns["type_mobile_app"], "FALSE")

    def test_inventory_types_from_categories(self):
        self.assertEqual(
            scanner.inventory_types_from_categories(["web_backend", "api_service", "middleware"]),
            ["web_app", "api_service", "middleware"],
        )

    def test_identifier_status(self):
        self.assertEqual(scanner.identifier_status("com.fabrikam.agsnap"), "found")
        self.assertEqual(scanner.identifier_status(""), "missing_from_scanned_files")

    def test_branch_name_from_ref(self):
        self.assertEqual(scanner.branch_name_from_ref("refs/heads/release/1.0"), "release/1.0")
        self.assertEqual(scanner.branch_name_from_ref("main"), "main")

    def test_default_branch_name_from_repo(self):
        self.assertEqual(scanner.default_branch_name_from_repo({"defaultBranch": "refs/heads/master"}), "master")
        self.assertEqual(scanner.default_branch_name_from_repo({"defaultBranch": "refs/heads/develop"}), "develop")
        self.assertEqual(scanner.default_branch_name_from_repo({}), "")

    def test_list_branch_targets_uses_only_default_branch(self):
        client = self.FakeBranchClient()
        target = scanner.RepoScanTarget(
            project_name="Project",
            repo={
                "id": "repo-id",
                "name": "Repo",
                "defaultBranch": "refs/heads/release",
            },
        )

        branch_targets = scanner.list_branch_targets(client, target)

        self.assertEqual(len(branch_targets), 1)
        self.assertEqual(branch_targets[0].branch_name, "release")
        self.assertEqual(client.branch_calls, 0)

    def test_select_fallback_branch_name_prefers_deployment_names(self):
        self.assertEqual(
            scanner.select_fallback_branch_name(["feature/foo", "main", "release/prod"]),
            "release/prod",
        )
        self.assertEqual(
            scanner.select_fallback_branch_name(["feature/foo", "development", "main"]),
            "main",
        )

    def test_list_branch_targets_uses_pipeline_branch_when_default_missing(self):
        client = self.FakeBranchClient(
            refs=[
                {"name": "refs/heads/main"},
                {"name": "refs/heads/release/prod"},
                {"name": "refs/heads/development"},
            ],
            definitions=[
                {
                    "repository": {"defaultBranch": "refs/heads/release/prod"},
                    "triggers": [{"branchFilters": ["+refs/heads/release/prod"]}],
                }
            ],
        )
        target = scanner.RepoScanTarget(
            project_name="Project",
            repo={
                "id": "repo-id",
                "name": "Repo",
            },
        )

        branch_targets = scanner.list_branch_targets(client, target)

        self.assertEqual(len(branch_targets), 1)
        self.assertEqual(branch_targets[0].branch_name, "release/prod")
        self.assertEqual(client.branch_calls, 1)
        self.assertEqual(client.definition_calls, 1)

    def test_list_branch_targets_uses_keyword_branch_when_default_missing(self):
        client = self.FakeBranchClient(
            refs=[
                {"name": "refs/heads/feature/foo"},
                {"name": "refs/heads/development"},
                {"name": "refs/heads/preprod"},
            ]
        )
        target = scanner.RepoScanTarget(
            project_name="Project",
            repo={
                "id": "repo-id",
                "name": "Repo",
            },
        )

        branch_targets = scanner.list_branch_targets(client, target)

        self.assertEqual(len(branch_targets), 1)
        self.assertEqual(branch_targets[0].branch_name, "preprod")

    def test_branch_age_bucket(self):
        now = datetime(2026, 6, 21, tzinfo=timezone.utc)

        self.assertEqual(scanner.branch_age_bucket("2026-06-01T00:00:00Z", 90, now), scanner.ACTIVE_SHEET_NAME)
        self.assertEqual(scanner.branch_age_bucket("2026-01-01T00:00:00Z", 90, now), scanner.OLDER_SHEET_NAME)
        self.assertEqual(scanner.branch_age_bucket("", 90, now), scanner.OLDER_SHEET_NAME)


class StoreLookupTests(unittest.TestCase):
    def test_target_store_platforms_uses_native_categories(self):
        self.assertEqual(scanner.target_store_platforms(["ios"]), (scanner.APPLE_PLATFORM,))
        self.assertEqual(scanner.target_store_platforms(["android"]), (scanner.GOOGLE_PLATFORM,))
        self.assertEqual(
            scanner.target_store_platforms(["react_native"]),
            (scanner.APPLE_PLATFORM, scanner.GOOGLE_PLATFORM),
        )

    def test_store_columns_from_found_listings(self):
        columns = scanner.store_columns_from_listings(
            [
                scanner.StoreListing(
                    platform=scanner.APPLE_PLATFORM,
                    status="found",
                    name="Agsnap",
                    identifier="com.fabrikam.agsnap",
                    url="https://apps.apple.com/app/id123",
                    version="1.0.2",
                    last_updated="2026-01-01T00:00:00Z",
                ),
                scanner.StoreListing(
                    platform=scanner.GOOGLE_PLATFORM,
                    status="not_found_publicly",
                    identifier="com.fabrikam.agsnap",
                ),
            ]
        )

        self.assertEqual(columns["store_lookup_status"], "partial_found")
        self.assertEqual(columns["store_validation_passed"], "FALSE")
        self.assertEqual(columns["store_platforms"], "Apple App Store")
        self.assertEqual(columns["apple_app_store_name"], "Agsnap")
        self.assertEqual(columns["apple_app_store_identifier"], "com.fabrikam.agsnap")
        self.assertEqual(columns["apple_app_store_validation_passed"], "TRUE")
        self.assertEqual(columns["google_play_validation_passed"], "FALSE")
        self.assertEqual(columns["google_play_lookup_status"], "not_found_publicly")

    def test_store_validation_passes_when_all_requested_stores_are_found(self):
        columns = scanner.store_columns_from_listings(
            [
                scanner.StoreListing(
                    platform=scanner.APPLE_PLATFORM,
                    status="found",
                    identifier="com.fabrikam.agsnap",
                ),
                scanner.StoreListing(
                    platform=scanner.GOOGLE_PLATFORM,
                    status="not_requested",
                ),
            ]
        )

        self.assertEqual(columns["store_validation_passed"], "TRUE")
        self.assertEqual(columns["apple_app_store_validation_passed"], "TRUE")
        self.assertEqual(columns["google_play_validation_passed"], "FALSE")

    def test_store_columns_disabled_and_identifier_missing(self):
        disabled = scanner.store_columns("com.fabrikam.agsnap", ["android"], None)
        missing = scanner.store_columns("", ["ios"], object())

        self.assertEqual(disabled["store_lookup_status"], "disabled")
        self.assertEqual(disabled["store_validation_passed"], "FALSE")
        self.assertEqual(disabled["google_play_lookup_status"], "disabled")
        self.assertEqual(disabled["google_play_validation_passed"], "FALSE")
        self.assertEqual(missing["store_lookup_status"], "identifier_missing")
        self.assertEqual(missing["store_validation_passed"], "FALSE")
        self.assertEqual(missing["apple_app_store_lookup_status"], "identifier_missing")

    def test_google_play_helpers(self):
        html = """\
<html>
<head>
  <title>Agsnap - Apps on Google Play</title>
  <meta property="og:title" content="Agsnap - Apps on Google Play" />
  <meta property="og:url" content="https://play.google.com/store/apps/details?id=com.fabrikam.agsnap" />
  <script>{"softwareVersion":"1.0.2","dateModified":"2026-01-02"}</script>
</head>
</html>
"""
        parser = scanner.MetaTagParser()
        parser.feed(html)

        self.assertEqual(scanner.normalize_google_play_title(parser.meta["og:title"]), "Agsnap")
        self.assertTrue(scanner.google_play_app_page(parser.meta, parser.meta["og:title"], "com.fabrikam.agsnap"))
        self.assertFalse(scanner.google_play_app_page({}, "Google Play", "com.fabrikam.agsnap"))
        self.assertEqual(scanner.extract_google_play_version(html), "1.0.2")
        self.assertEqual(scanner.extract_google_play_updated(html), "2026-01-02")


class CliTests(unittest.TestCase):
    def test_parse_args_requires_pat(self):
        with self.assertRaises(SystemExit):
            scanner.parse_args(["--org", "example"])

    def test_confidence_rank_rejects_unknown_values(self):
        with self.assertRaises(Exception):
            scanner.confidence_rank("banana")

    def test_parse_args_rejects_invalid_content_workers(self):
        with self.assertRaises(SystemExit):
            scanner.parse_args(
                [
                    "--org",
                    "example",
                    "--pat",
                    "token",
                    "--content-workers",
                    "0",
                ]
            )

    def test_parse_args_rejects_invalid_commit_limit(self):
        with self.assertRaises(SystemExit):
            scanner.parse_args(
                [
                    "--org",
                    "example",
                    "--pat",
                    "token",
                    "--max-commits-per-repo",
                    "-1",
                ]
            )

    def test_parse_args_rejects_invalid_branch_age_days(self):
        with self.assertRaises(SystemExit):
            scanner.parse_args(
                [
                    "--org",
                    "example",
                    "--pat",
                    "token",
                    "--branch-age-days",
                    "0",
                ]
            )

    def test_parse_args_accepts_store_lookup_options(self):
        config = scanner.parse_args(
            [
                "--org",
                "example",
                "--pat",
                "token",
                "--branch-workers",
                "24",
                "--activity-mode",
                "latest",
                "--store-lookup",
                "--store-country",
                "ca",
                "--store-timeout",
                "7",
            ]
        )

        self.assertEqual(config.branch_workers, 24)
        self.assertEqual(config.activity_mode, "latest")
        self.assertTrue(config.store_lookup)
        self.assertEqual(config.store_country, "CA")
        self.assertEqual(config.store_timeout_seconds, 7)

    def test_parse_args_rejects_invalid_store_options(self):
        with self.assertRaises(SystemExit):
            scanner.parse_args(["--org", "example", "--pat", "token", "--store-country", "usa"])
        with self.assertRaises(SystemExit):
            scanner.parse_args(["--org", "example", "--pat", "token", "--store-timeout", "0"])

    def test_parse_args_rejects_invalid_branch_workers(self):
        with self.assertRaises(SystemExit):
            scanner.parse_args(["--org", "example", "--pat", "token", "--branch-workers", "0"])


if __name__ == "__main__":
    unittest.main()
