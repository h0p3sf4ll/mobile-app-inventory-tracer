from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import TextIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .constants import (
    CSV_FIELDNAMES,
    DEFAULT_BRANCH_AGE_DAYS,
    SCANNER_TARGET_FIELDNAMES,
    SONARQUBE_FIELDNAMES,
    active_sheet_name,
    older_sheet_name,
)

WORKBOOK_COLUMN_WIDTHS = {
    "project": 24,
    "repo_name": 28,
    "branch_name": 24,
    "branch_last_updated": 22,
    "branch_age_bucket": 14,
    "web_url": 42,
    "source_url": 52,
    "inventory_name": 30,
    "inventory_version": 18,
    "inventory_types": 30,
    "primary_language": 18,
    "scanner_target": 60,
    "semgrep_target": 60,
    "sonarqube_project_key": 44,
    "sonarqube_project_name": 34,
    "mobile_name": 28,
    "mobile_version": 16,
    "mobile_identifier": 34,
    "mobile_identifier_source": 28,
    "mobile_identifier_status": 24,
    "contributing_developers": 70,
    "last_updated": 22,
    "confidence": 12,
    "score": 10,
    "categories": 34,
    "store_lookup_status": 18,
    "store_validation_passed": 24,
    "store_platforms": 30,
    "apple_app_store_name": 30,
    "apple_app_store_identifier": 34,
    "apple_app_store_url": 52,
    "apple_app_store_version": 18,
    "apple_app_store_last_updated": 24,
    "apple_app_store_validation_passed": 32,
    "apple_app_store_lookup_status": 26,
    "google_play_name": 30,
    "google_play_identifier": 34,
    "google_play_url": 52,
    "google_play_version": 18,
    "google_play_last_updated": 24,
    "google_play_validation_passed": 30,
    "google_play_lookup_status": 24,
    "detection_evidence": 80,
}


def write_outputs(
    results: list[dict[str, object]],
    out_dir: Path,
    out_prefix: str,
    branch_age_days: int = DEFAULT_BRANCH_AGE_DAYS,
) -> tuple[Path, Path, Path]:
    with StreamingReportWriter(out_dir, out_prefix, branch_age_days) as writer:
        for result in results:
            writer.write_result(result)
        return writer.csv_path, writer.json_path, writer.xlsx_path


class StreamingReportWriter:
    def __init__(self, out_dir: Path, out_prefix: str, branch_age_days: int = DEFAULT_BRANCH_AGE_DAYS) -> None:
        self.out_dir = out_dir
        self.out_prefix = out_prefix
        self.active_sheet_name = active_sheet_name(branch_age_days)
        self.older_sheet_name = older_sheet_name(branch_age_days)
        self.csv_path = out_dir / f"{out_prefix}.csv"
        self.json_path = out_dir / f"{out_prefix}.json"
        self.xlsx_path = out_dir / f"{out_prefix}.xlsx"
        self.scanner_targets_csv_path = out_dir / f"{out_prefix}_scanner_targets.csv"
        self.scanner_targets_json_path = out_dir / f"{out_prefix}_scanner_targets.json"
        self.semgrep_targets_path = out_dir / f"{out_prefix}_semgrep_targets.txt"
        self.sonarqube_projects_path = out_dir / f"{out_prefix}_sonarqube_projects.csv"
        self._csv_file: TextIO | None = None
        self._json_file: TextIO | None = None
        self._scanner_targets_csv_file: TextIO | None = None
        self._scanner_targets_json_file: TextIO | None = None
        self._semgrep_targets_file: TextIO | None = None
        self._sonarqube_projects_file: TextIO | None = None
        self._csv_writer: csv.DictWriter[str] | None = None
        self._scanner_targets_writer: csv.DictWriter[str] | None = None
        self._sonarqube_projects_writer: csv.DictWriter[str] | None = None
        self._workbook: Workbook | None = None
        self._sheets: dict[str, Worksheet] = {}
        self._row_count = 0
        self._xlsx_save_interval = 25

    def __enter__(self) -> "StreamingReportWriter":
        self.out_dir.mkdir(parents=True, exist_ok=True)
        self._csv_file = self.csv_path.open("w", newline="", encoding="utf-8")
        self._json_file = self.json_path.open("w", encoding="utf-8")
        self._scanner_targets_csv_file = self.scanner_targets_csv_path.open("w", newline="", encoding="utf-8")
        self._scanner_targets_json_file = self.scanner_targets_json_path.open("w", encoding="utf-8")
        self._semgrep_targets_file = self.semgrep_targets_path.open("w", encoding="utf-8")
        self._sonarqube_projects_file = self.sonarqube_projects_path.open("w", newline="", encoding="utf-8")
        self._csv_writer = csv.DictWriter(self._csv_file, fieldnames=CSV_FIELDNAMES)
        self._scanner_targets_writer = csv.DictWriter(
            self._scanner_targets_csv_file,
            fieldnames=SCANNER_TARGET_FIELDNAMES,
        )
        self._sonarqube_projects_writer = csv.DictWriter(
            self._sonarqube_projects_file,
            fieldnames=SONARQUBE_FIELDNAMES,
        )
        self._csv_writer.writeheader()
        self._scanner_targets_writer.writeheader()
        self._sonarqube_projects_writer.writeheader()
        self._json_file.write("[\n")
        self._scanner_targets_json_file.write("[\n")
        self._create_workbook()
        self._save_workbook()
        self.flush()
        return self

    def __exit__(self, exc_type: object, exc_value: object, traceback: object) -> None:
        self.close()

    def write_result(self, result: dict[str, object]) -> None:
        if (
            self._csv_writer is None
            or self._json_file is None
            or self._scanner_targets_writer is None
            or self._scanner_targets_json_file is None
            or self._sonarqube_projects_writer is None
        ):
            raise RuntimeError("StreamingReportWriter must be opened before writing.")

        self._csv_writer.writerow(result)
        target_row = scanner_target_row(result)
        self._scanner_targets_writer.writerow(target_row)
        self._sonarqube_projects_writer.writerow(sonarqube_project_row(result))
        if self._semgrep_targets_file and result.get("semgrep_target"):
            self._semgrep_targets_file.write(f"{result.get('semgrep_target')}\n")
        if self._row_count:
            self._json_file.write(",\n")
            self._scanner_targets_json_file.write(",\n")
        json.dump(result, self._json_file, indent=2)
        json.dump(target_row, self._scanner_targets_json_file, indent=2)
        self._row_count += 1
        self._append_workbook_row(result)
        if self._row_count % self._xlsx_save_interval == 0:
            self._save_workbook()
        self.flush()

    def flush(self) -> None:
        if self._csv_file:
            self._csv_file.flush()
        if self._json_file:
            self._json_file.flush()
        if self._scanner_targets_csv_file:
            self._scanner_targets_csv_file.flush()
        if self._scanner_targets_json_file:
            self._scanner_targets_json_file.flush()
        if self._semgrep_targets_file:
            self._semgrep_targets_file.flush()
        if self._sonarqube_projects_file:
            self._sonarqube_projects_file.flush()

    def close(self) -> None:
        self._save_workbook()
        if self._json_file:
            self._json_file.write("\n]\n")
            self._json_file.close()
            self._json_file = None
        if self._scanner_targets_json_file:
            self._scanner_targets_json_file.write("\n]\n")
            self._scanner_targets_json_file.close()
            self._scanner_targets_json_file = None
        if self._csv_file:
            self._csv_file.close()
            self._csv_file = None
        if self._scanner_targets_csv_file:
            self._scanner_targets_csv_file.close()
            self._scanner_targets_csv_file = None
        if self._semgrep_targets_file:
            self._semgrep_targets_file.close()
            self._semgrep_targets_file = None
        if self._sonarqube_projects_file:
            self._sonarqube_projects_file.close()
            self._sonarqube_projects_file = None
        self._csv_writer = None
        self._scanner_targets_writer = None
        self._sonarqube_projects_writer = None
        if self._workbook:
            self._workbook.close()
            self._workbook = None
            self._sheets = {}

    def _create_workbook(self) -> None:
        self._workbook = Workbook()
        active_sheet = self._workbook.active
        active_sheet.title = self.active_sheet_name
        older_sheet = self._workbook.create_sheet(self.older_sheet_name)
        self._sheets = {
            self.active_sheet_name: active_sheet,
            self.older_sheet_name: older_sheet,
        }
        for sheet in self._sheets.values():
            sheet.append(list(CSV_FIELDNAMES))
            sheet.freeze_panes = "A2"
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
            self._apply_column_widths(sheet)

    def _append_workbook_row(self, result: dict[str, object]) -> None:
        if self._workbook is None:
            return
        sheet = self._sheets.get(result.get("branch_age_bucket")) or self._sheets[self.older_sheet_name]
        sheet.append([result.get(field, "") for field in CSV_FIELDNAMES])

    def _save_workbook(self) -> None:
        if self._workbook is None:
            return
        for sheet in self._sheets.values():
            sheet.auto_filter.ref = sheet.dimensions
        self._workbook.save(self.xlsx_path)

    @staticmethod
    def _apply_column_widths(sheet: Worksheet) -> None:
        for index, field_name in enumerate(CSV_FIELDNAMES, start=1):
            column_letter = get_column_letter(index)
            sheet.column_dimensions[column_letter].width = WORKBOOK_COLUMN_WIDTHS.get(field_name, 16)


def scanner_target_row(result: dict[str, object]) -> dict[str, object]:
    return {field_name: result.get(field_name, "") for field_name in SCANNER_TARGET_FIELDNAMES}


def sonarqube_project_row(result: dict[str, object]) -> dict[str, object]:
    return {
        "sonar.projectKey": result.get("sonarqube_project_key", ""),
        "sonar.projectName": result.get("sonarqube_project_name", ""),
        "sonar.sources": ".",
        "branch": result.get("branch_name", ""),
        "source_url": result.get("source_url", ""),
        "web_url": result.get("web_url", ""),
        "project": result.get("project", ""),
        "repo_name": result.get("repo_name", ""),
        "inventory_types": result.get("inventory_types", ""),
        "categories": result.get("categories", ""),
    }
